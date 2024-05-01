#change the file_path and sheet_name in main function according to the excel sheet to run the code 
import openpyxl
import datetime

e = datetime.datetime.now()
date=e.strftime("%d%m%Y")

def get_data_from_excel_cell(file_path, sheet_name, cell_address):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    cell_data = sheet[cell_address].value

    if cell_data is not None:
        data_list = cell_data.split(',')
        return data_list
    else:
        return []

def write_to_file(file_name, data_list):
    with open(file_name, 'w') as file:
        for data in data_list:
            file.write(data + '\n')

def process_row_data(b2_cell_data, c2_cell_data, e2_cell_data, f2_cell_data):
    concatenated_data_list = []
    table_data_list = []
    index_data_list = []

    if e2_cell_data:
        data_list = e2_cell_data.split(',')
        for item in data_list:
            concatenated_row_data = f'{b2_cell_data}.{c2_cell_data}:{item.strip()}, '
            concatenated_data_list.append(concatenated_row_data)
            table_data_list.append(f'alter table {b2_cell_data}.{c2_cell_data} drop partition {item.strip()} update global indexes;')

    if f2_cell_data:
        data_list = f2_cell_data.split(',')
        for item in data_list:
            index_data_list.append(f'alter index {b2_cell_data}.{item.strip()} rebuild online parallel 10;')

    return concatenated_data_list, table_data_list, index_data_list

def main():
    file_path = 'MainScript.xlsx'
    sheet_name = 'Testing'
    cell_address = 'E2'


    result_list = get_data_from_excel_cell(file_path, sheet_name, cell_address)

    # Process and write data to the file for each row
    with open('output.txt', 'w') as file:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            b2_cell_data, c2_cell_data, e2_cell_data = row[1], row[2], row[4]
            f2_cell_data = row[5]
            d2_cell_data = row[3]


            file.write(
                f'expdp directory=DPUMP1 dumpfile=expdp_{c2_cell_data}_{d2_cell_data}_{date}.dmp logfile=expdp_{c2_cell_data}_{d2_cell_data}_{date}.log tables=')


            concatenated_data_list, table_data_list, index_data_list = process_row_data(b2_cell_data, c2_cell_data,
                                                                                        e2_cell_data, f2_cell_data)


            for data in concatenated_data_list:
                file.write(data)
            file.write('\n\n')


            for data in table_data_list:
                file.write(data + '\n')
            file.write('\n')

            for data in index_data_list:
                file.write(data + '\n')
            file.write('\n\n')  #

if __name__ == "__main__":
    main()
